"""Document chunking and text extraction."""

import io
import csv
import gc
import logging
from typing import List

log = logging.getLogger("localisa.rag.chunker")


def chunk_text(text: str, chunk_size: int = 800, chunk_overlap: int = 150, min_size: int = 50) -> List[str]:
    """Split text into overlapping chunks on natural boundaries."""
    text = text.strip()
    if not text:
        return []
    if len(text) <= chunk_size:
        return [text] if len(text) >= min_size else []

    chunks = []
    start = 0
    max_iter = len(text) // 100 + 10
    i = 0

    while start < len(text) and i < max_iter:
        i += 1
        end = min(start + chunk_size, len(text))

        if end < len(text):
            for sep in ['\n\n', '\n', '. ', ', ']:
                last_sep = text.rfind(sep, start + min_size, end)
                if last_sep > start:
                    end = last_sep + len(sep)
                    break

        chunk = text[start:end].strip()
        if chunk and len(chunk) >= min_size:
            chunks.append(chunk)

        new_start = end - chunk_overlap
        if new_start <= start:
            new_start = start + chunk_size // 2
        start = new_start

        if start >= len(text) - min_size:
            remaining = text[start:].strip()
            if remaining and len(remaining) >= min_size and remaining not in chunks:
                chunks.append(remaining)
            break

    return chunks


def extract_text(filename: str, content: bytes) -> str:
    """Extract text from various file formats."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    text = ""

    try:
        if ext == 'docx':
            import docx
            doc = docx.Document(io.BytesIO(content))
            text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])

        elif ext == 'pdf':
            import fitz
            doc = fitz.open(stream=content, filetype="pdf")
            text = '\n'.join([page.get_text() for page in doc])
            doc.close()

        elif ext == 'xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                sheet = wb[name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else '' for c in row]
                    if any(cells):
                        rows.append(' | '.join(cells))
                if rows:
                    sheets.append(f"[Sheet: {name}]\n" + '\n'.join(rows))
            text = '\n\n'.join(sheets)
            wb.close()

        elif ext in ('txt', 'md', 'csv'):
            text = content.decode('utf-8', errors='ignore')

        else:
            text = content.decode('utf-8', errors='ignore')

    except Exception as e:
        log.error(f"Error extracting text from {filename}: {e}")
    finally:
        gc.collect()

    return text
